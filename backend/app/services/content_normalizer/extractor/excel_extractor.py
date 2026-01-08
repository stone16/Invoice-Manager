from __future__ import annotations

from io import BytesIO
from typing import List, Set

import openpyxl
from openpyxl.utils import get_column_letter

from app.models.digi_flow import FileContentType
from app.services.content_normalizer.block_id import (
    build_excel_block_id,
    build_merged_cell_block_id,
)
from app.services.content_normalizer.models import BoundingBox, FileContentMetadata, Sheet, SheetContent


class ExcelExtractor:
    def extract(
        self,
        file_bytes: bytes,
        doc_index: int,
        file_name: str,
        file_object_fid: str,
    ) -> FileContentMetadata:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        sheets: List[Sheet] = []

        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            bounding_boxes: List[BoundingBox] = []
            merged_ranges = list(sheet.merged_cells.ranges)
            merged_cells: Set[str] = set()
            for merged_range in merged_ranges:
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_cells.add(f"{get_column_letter(col)}{row}")

            for merged_range in merged_ranges:
                start_cell = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
                cell = sheet[start_cell]
                value = cell.value
                if value is None or str(value).strip() == "":
                    continue
                bbox = BoundingBox(
                    id=build_merged_cell_block_id(
                        doc_index,
                        sheet_index,
                        f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}",
                        f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}",
                    ),
                    raw_value=str(value),
                    top_left_x=float(merged_range.min_col),
                    top_left_y=float(merged_range.min_row),
                    bottom_right_x=float(merged_range.max_col),
                    bottom_right_y=float(merged_range.max_row),
                )
                bounding_boxes.append(bbox)

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.coordinate in merged_cells:
                        continue
                    value = cell.value
                    if value is None or str(value).strip() == "":
                        continue
                    bbox = BoundingBox(
                        id=build_excel_block_id(doc_index, sheet_index, cell.coordinate),
                        raw_value=str(value),
                        top_left_x=float(cell.column),
                        top_left_y=float(cell.row),
                        bottom_right_x=float(cell.column + 1),
                        bottom_right_y=float(cell.row + 1),
                    )
                    bounding_boxes.append(bbox)

            sheets.append(Sheet(id=sheet_index, name=sheet.title, bounding_boxes=bounding_boxes))

        content = SheetContent(sheets=sheets)
        return FileContentMetadata(
            index=doc_index,
            file_object_fid=file_object_fid,
            file_name=file_name,
            file_bytes_size=len(file_bytes),
            content_type=FileContentType.EXCEL,
            languages=["zh"],
            file_content=content,
        )
