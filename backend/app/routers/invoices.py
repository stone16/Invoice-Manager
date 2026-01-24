from typing import Optional, List
from io import BytesIO
from datetime import date
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, BackgroundTasks, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from decimal import Decimal

from app.database import get_db
from app.models.invoice import Invoice, OcrResult, LlmResult, ParsingDiff, InvoiceStatus
from app.schemas.invoice import (
    InvoiceResponse, InvoiceListResponse, InvoiceDetailResponse,
    InvoiceUpdate, BatchUpdateRequest, BatchDeleteRequest, StatisticsResponse, UploadResponse,
    ResolveDiffRequest
)
from app.config import get_settings
from app.services.audit_service import log_audit_no_commit, get_client_info
from app.rate_limit import limiter

settings = get_settings()
router = APIRouter()


def _parse_invoice_ids(invoice_ids: Optional[str]) -> Optional[List[int]]:
    if not invoice_ids:
        return None
    ids: List[int] = []
    for raw_id in invoice_ids.split(","):
        raw_id = raw_id.strip()
        if not raw_id:
            continue
        try:
            ids.append(int(raw_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="发票ID必须为整数") from exc
    if not ids:
        raise HTTPException(status_code=400, detail="发票ID不能为空")
    return ids


def _parse_date_param(value: Optional[str], field_name: str) -> Optional[date]:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} 日期格式无效，应为 YYYY-MM-DD") from exc


@router.post("/upload", response_model=List[UploadResponse])
@limiter.limit("10/minute")
async def upload_invoices(
    request: Request,
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db)
):
    """上传发票文件 (支持多文件)，上传后异步触发OCR解析"""
    results = []
    invoice_ids_to_process = []
    client_info = get_client_info(request)

    for file in files:
        # Validate file type
        ext = file.filename.split(".")[-1].lower() if file.filename else ""
        if ext not in settings.allowed_extensions:
            results.append(UploadResponse(
                id=0,
                file_name=file.filename or "unknown",
                status="error",
                message=f"不支持的文件类型: {ext}"
            ))
            continue

        # Read file content
        content = await file.read()

        # Validate file size
        if len(content) > settings.max_file_size:
            results.append(UploadResponse(
                id=0,
                file_name=file.filename or "unknown",
                status="error",
                message=f"文件过大，最大支持 {settings.max_file_size // 1024 // 1024}MB"
            ))
            continue

        # Create invoice record with UPLOADED status (not yet processed)
        invoice = Invoice(
            file_name=file.filename or "unknown",
            file_type=ext,
            file_data=content,
            status=InvoiceStatus.UPLOADED
        )
        db.add(invoice)
        await db.flush()

        invoice_ids_to_process.append(invoice.id)

        # Audit log for upload
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="upload",
            new_value={"file_name": invoice.file_name, "file_type": ext, "file_size": len(content)},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )

        results.append(UploadResponse(
            id=invoice.id,
            file_name=invoice.file_name,
            status="success",
            message="上传成功，等待解析"
        ))

    await db.commit()

    # Schedule background processing for each uploaded invoice
    for invoice_id in invoice_ids_to_process:
        background_tasks.add_task(process_invoice_background, invoice_id)

    return results


async def process_invoice_background(invoice_id: int, max_retries: int = 3):
    """Background task to process an invoice with OCR/LLM.

    Implements exponential backoff retry logic for transient failures.
    Total attempts = 1 (initial) + max_retries, with delays of 2, 4, 8 seconds
    between retries.

    Args:
        invoice_id: ID of the invoice to process
        max_retries: Maximum number of retries after the initial attempt (default: 3,
            resulting in up to 4 total attempts)
    """
    from app.services.invoice_service import process_invoice as do_process
    from app.database import async_session_maker
    from app.services.audit_service import log_audit
    import logging
    import asyncio
    logger = logging.getLogger(__name__)

    retry_count = 0
    last_error = None

    while retry_count <= max_retries:
        async with async_session_maker() as db:
            try:
                # Update status to PROCESSING
                query = select(Invoice).where(Invoice.id == invoice_id)
                result = await db.execute(query)
                invoice = result.scalar_one_or_none()

                if not invoice:
                    logger.error(f"Invoice {invoice_id} not found")
                    return

                invoice.status = InvoiceStatus.PROCESSING
                await db.commit()

                logger.info(f"Background processing invoice {invoice_id} (attempt {retry_count + 1}/{max_retries + 1})")
                success = await do_process(invoice_id, db)

                if success:
                    logger.info(f"Invoice {invoice_id} processing completed successfully")
                    # Log successful processing
                    await log_audit(
                        db=db,
                        entity_type="invoice",
                        entity_id=invoice_id,
                        action="process_complete",
                        new_value={"status": "success", "attempts": retry_count + 1}
                    )
                    return
                else:
                    last_error = "Processing returned false"
                    logger.warning(f"Invoice {invoice_id} processing returned false")

            except Exception as e:
                last_error = str(e)
                logger.error(f"Failed to process invoice {invoice_id} (attempt {retry_count + 1}): {e}")
                await db.rollback()

        retry_count += 1

        if retry_count <= max_retries:
            # Exponential backoff: 2^retry_count seconds (2, 4, 8 seconds)
            delay = 2 ** retry_count
            logger.info(f"Retrying invoice {invoice_id} in {delay} seconds...")
            await asyncio.sleep(delay)

    # All retries exhausted - mark as failed
    async with async_session_maker() as db:
        try:
            query = select(Invoice).where(Invoice.id == invoice_id)
            result = await db.execute(query)
            invoice = result.scalar_one_or_none()

            if invoice:
                # Set status back to UPLOADED so user can retry manually
                invoice.status = InvoiceStatus.UPLOADED
                await db.commit()

                # Log failed processing
                await log_audit(
                    db=db,
                    entity_type="invoice",
                    entity_id=invoice_id,
                    action="process_failed",
                    new_value={"error": last_error, "attempts": max_retries + 1}
                )

            logger.error(f"Invoice {invoice_id} processing failed after {max_retries + 1} attempts: {last_error}")
        except Exception as e:
            logger.error(f"Failed to update invoice {invoice_id} status after retry exhaustion: {e}")


@router.get("", response_model=InvoiceListResponse)
async def list_invoices(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
    owner: Optional[str] = Query(None, description="归属人筛选"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db)
):
    """获取发票列表"""
    query = select(Invoice)

    # Apply filters
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    if start_date:
        try:
            start_date_obj = date.fromisoformat(start_date)
            query = query.where(Invoice.issue_date >= start_date_obj)
        except ValueError:
            pass  # Invalid date format, skip filter
    if end_date:
        try:
            end_date_obj = date.fromisoformat(end_date)
            query = query.where(Invoice.issue_date <= end_date_obj)
        except ValueError:
            pass  # Invalid date format, skip filter

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total = await db.scalar(count_query) or 0

    # Apply pagination
    query = query.order_by(Invoice.created_at.desc())
    query = query.offset((page - 1) * page_size).limit(page_size)

    result = await db.execute(query)
    invoices = result.scalars().all()

    return InvoiceListResponse(
        items=[InvoiceResponse.model_validate(inv) for inv in invoices],
        total=total,
        page=page,
        page_size=page_size
    )


@router.get("/statistics", response_model=StatisticsResponse)
async def get_statistics(
    invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
    status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
    owner: Optional[str] = Query(None, description="归属人筛选"),
    db: AsyncSession = Depends(get_db)
):
    """获取发票统计数据"""
    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)

    result = await db.execute(query)
    invoices = result.scalars().all()

    count = len(invoices)
    total_amount = sum((inv.amount or Decimal(0)) for inv in invoices)
    total_tax = sum((inv.tax_amount or Decimal(0)) for inv in invoices)
    total_with_tax = sum((inv.total_with_tax or Decimal(0)) for inv in invoices)

    return StatisticsResponse(
        count=count,
        total_amount=total_amount,
        total_tax=total_tax,
        total_with_tax=total_with_tax
    )


@router.get("/{invoice_id}", response_model=InvoiceDetailResponse)
async def get_invoice(
    invoice_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取发票详情"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Load related data
    ocr_query = select(OcrResult).where(OcrResult.invoice_id == invoice_id)
    ocr_result = await db.execute(ocr_query)
    ocr = ocr_result.scalar_one_or_none()

    llm_query = select(LlmResult).where(LlmResult.invoice_id == invoice_id)
    llm_result = await db.execute(llm_query)
    llm = llm_result.scalar_one_or_none()

    diff_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    diff_result = await db.execute(diff_query)
    diffs = diff_result.scalars().all()

    # Build response dict from invoice to avoid lazy loading issues
    invoice_dict = {
        "id": invoice.id,
        "file_name": invoice.file_name,
        "file_type": invoice.file_type,
        "status": invoice.status,
        "owner": invoice.owner,
        "invoice_number": invoice.invoice_number,
        "issue_date": invoice.issue_date,
        "buyer_name": invoice.buyer_name,
        "buyer_tax_id": invoice.buyer_tax_id,
        "seller_name": invoice.seller_name,
        "seller_tax_id": invoice.seller_tax_id,
        "item_name": invoice.item_name,
        "total_with_tax": invoice.total_with_tax,
        "specification": invoice.specification,
        "unit": invoice.unit,
        "quantity": invoice.quantity,
        "unit_price": invoice.unit_price,
        "amount": invoice.amount,
        "tax_rate": invoice.tax_rate,
        "tax_amount": invoice.tax_amount,
        "created_at": invoice.created_at,
        "updated_at": invoice.updated_at,
        "ocr_result": ocr,
        "llm_result": llm,
        "parsing_diffs": list(diffs),
    }

    return InvoiceDetailResponse.model_validate(invoice_dict)


@router.get("/{invoice_id}/file")
async def get_invoice_file(
    invoice_id: int,
    db: AsyncSession = Depends(get_db)
):
    """获取发票原始文件"""
    from fastapi.responses import Response
    from urllib.parse import quote

    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    content_type_map = {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png"
    }

    # URL-encode filename for Content-Disposition header (RFC 5987)
    encoded_filename = quote(invoice.file_name)

    return Response(
        content=invoice.file_data,
        media_type=content_type_map.get(invoice.file_type, "application/octet-stream"),
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.put("/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(
    invoice_id: int,
    update_data: InvoiceUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """更新发票信息"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Capture old values for audit
    update_dict = update_data.model_dump(exclude_unset=True)
    old_values = {key: getattr(invoice, key) for key in update_dict.keys()}
    # Convert non-serializable types
    for key, value in old_values.items():
        if hasattr(value, 'value'):  # Enum
            old_values[key] = value.value
        elif hasattr(value, 'isoformat'):  # Date/DateTime
            old_values[key] = value.isoformat()
        elif isinstance(value, Decimal):
            old_values[key] = str(value)

    for key, value in update_dict.items():
        setattr(invoice, key, value)

    # Audit log
    client_info = get_client_info(request)
    new_values = update_dict.copy()
    for key, value in new_values.items():
        if hasattr(value, 'value'):  # Enum
            new_values[key] = value.value
        elif hasattr(value, 'isoformat'):  # Date/DateTime
            new_values[key] = value.isoformat()
        elif isinstance(value, Decimal):
            new_values[key] = str(value)

    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="update",
        old_value=old_values,
        new_value=new_values,
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()
    await db.refresh(invoice)

    return InvoiceResponse.model_validate(invoice)


@router.post("/batch-update")
@limiter.limit("30/minute")
async def batch_update_invoices(
    request: Request,
    batch_request: BatchUpdateRequest,
    db: AsyncSession = Depends(get_db)
):
    """批量更新发票状态/归属人"""
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    client_info = get_client_info(request)
    updated_count = 0
    for invoice in invoices:
        old_values = {}
        new_values = {}
        if batch_request.status is not None:
            old_values["status"] = invoice.status.value if invoice.status else None
            invoice.status = batch_request.status
            new_values["status"] = batch_request.status.value
        if batch_request.owner is not None:
            old_values["owner"] = invoice.owner
            invoice.owner = batch_request.owner
            new_values["owner"] = batch_request.owner
        updated_count += 1

        # Audit log for each invoice
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="batch_update",
            old_value=old_values,
            new_value=new_values,
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )

    await db.commit()

    return {
        "message": f"成功更新 {updated_count} 张发票",
        "updated_count": updated_count
    }


@router.post("/batch-delete")
@limiter.limit("20/minute")
async def batch_delete_invoices(
    request: Request,
    batch_request: BatchDeleteRequest,
    db: AsyncSession = Depends(get_db)
):
    """批量删除发票及其关联数据"""
    if not batch_request.invoice_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的发票")

    # Query invoices to delete
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(status_code=404, detail="未找到要删除的发票")

    client_info = get_client_info(request)
    deleted_count = 0
    for invoice in invoices:
        # Audit log for each deletion
        await log_audit_no_commit(
            db=db,
            entity_type="invoice",
            entity_id=invoice.id,
            action="delete",
            old_value={"file_name": invoice.file_name, "invoice_number": invoice.invoice_number},
            ip_address=client_info.get("ip_address"),
            user_agent=client_info.get("user_agent"),
        )
        await db.delete(invoice)
        deleted_count += 1

    await db.commit()

    return {
        "message": f"成功删除 {deleted_count} 张发票",
        "deleted_count": deleted_count
    }


@router.post("/batch-reprocess")
@limiter.limit("5/minute")
async def batch_reprocess_invoices(
    request: Request,
    background_tasks: BackgroundTasks,
    batch_request: BatchDeleteRequest,  # Reuse for invoice_ids
    db: AsyncSession = Depends(get_db)
):
    """批量重新解析发票（清除旧的OCR/LLM结果，重新处理）"""
    import logging
    logger = logging.getLogger(__name__)

    if not batch_request.invoice_ids:
        raise HTTPException(status_code=400, detail="请选择要重新解析的发票")

    # Query invoices to reprocess
    query = select(Invoice).where(Invoice.id.in_(batch_request.invoice_ids))
    result = await db.execute(query)
    invoices = result.scalars().all()

    if not invoices:
        raise HTTPException(status_code=404, detail="未找到要重新解析的发票")

    # Clear old parsing results and reset invoice fields
    for invoice in invoices:
        # Delete old OCR results
        ocr_query = select(OcrResult).where(OcrResult.invoice_id == invoice.id)
        ocr_result = await db.execute(ocr_query)
        for ocr in ocr_result.scalars().all():
            await db.delete(ocr)

        # Delete old LLM results
        llm_query = select(LlmResult).where(LlmResult.invoice_id == invoice.id)
        llm_result = await db.execute(llm_query)
        for llm in llm_result.scalars().all():
            await db.delete(llm)

        # Delete old parsing diffs
        diff_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice.id)
        diff_result = await db.execute(diff_query)
        for diff in diff_result.scalars().all():
            await db.delete(diff)

        # Reset invoice fields
        invoice.invoice_number = None
        invoice.issue_date = None
        invoice.buyer_name = None
        invoice.buyer_tax_id = None
        invoice.seller_name = None
        invoice.seller_tax_id = None
        invoice.item_name = None
        invoice.total_with_tax = None
        invoice.amount = None
        invoice.tax_amount = None
        invoice.tax_rate = None
        invoice.status = InvoiceStatus.UPLOADED

    await db.commit()
    logger.info(f"Cleared old parsing results for {len(invoices)} invoices, scheduling reprocess")

    # Schedule background processing
    for invoice in invoices:
        background_tasks.add_task(process_invoice_background, invoice.id)

    return {
        "message": f"已清除 {len(invoices)} 张发票的旧解析结果，正在重新解析",
        "count": len(invoices)
    }


@router.post("/{invoice_id}/process")
async def process_invoice(
    invoice_id: int,
    db: AsyncSession = Depends(get_db)
):
    """处理发票：运行OCR解析"""
    from app.services.invoice_service import process_invoice as do_process

    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    success = await do_process(invoice_id, db)

    if success:
        return {"message": "解析成功", "invoice_id": invoice_id}
    else:
        raise HTTPException(status_code=500, detail="解析失败")


@router.delete("/{invoice_id}")
async def delete_invoice(
    invoice_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """删除发票"""
    query = select(Invoice).where(Invoice.id == invoice_id)
    result = await db.execute(query)
    invoice = result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Audit log
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="delete",
        old_value={"file_name": invoice.file_name, "invoice_number": invoice.invoice_number},
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.delete(invoice)
    await db.commit()

    return {"message": "删除成功"}


@router.post("/{invoice_id}/diffs/{diff_id}/resolve")
async def resolve_diff(
    invoice_id: int,
    diff_id: int,
    resolve_request: ResolveDiffRequest,
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """解决解析差异，选择OCR、LLM或自定义值"""
    from datetime import datetime
    from decimal import Decimal as Dec

    # Get the diff
    diff_query = select(ParsingDiff).where(
        ParsingDiff.id == diff_id,
        ParsingDiff.invoice_id == invoice_id
    )
    diff_result = await db.execute(diff_query)
    diff = diff_result.scalar_one_or_none()

    if not diff:
        raise HTTPException(status_code=404, detail="差异记录不存在")

    # Capture old value for audit
    old_diff_value = {
        "field_name": diff.field_name,
        "final_value": diff.final_value,
        "source": diff.source,
        "resolved": diff.resolved,
    }

    # Determine the final value
    if resolve_request.source == 'ocr':
        final_value = diff.ocr_value
    elif resolve_request.source == 'llm':
        final_value = diff.llm_value
    elif resolve_request.source == 'custom':
        final_value = resolve_request.custom_value
    else:
        raise HTTPException(status_code=400, detail="无效的来源类型")

    # Update the diff
    diff.final_value = final_value
    diff.source = resolve_request.source
    diff.resolved = 1

    # Get the invoice and update the corresponding field
    invoice_query = select(Invoice).where(Invoice.id == invoice_id)
    invoice_result = await db.execute(invoice_query)
    invoice = invoice_result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    # Update the invoice field based on field_name
    field_name = diff.field_name
    if final_value is not None:
        if field_name == 'issue_date':
            try:
                invoice.issue_date = datetime.strptime(final_value, '%Y-%m-%d').date()
            except ValueError:
                pass
        elif field_name in ['total_with_tax', 'amount', 'tax_amount', 'quantity', 'unit_price']:
            try:
                setattr(invoice, field_name, Dec(final_value))
            except (ValueError, TypeError):
                pass
        else:
            setattr(invoice, field_name, final_value)

    # Check if all diffs are resolved
    all_diffs_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    all_diffs_result = await db.execute(all_diffs_query)
    all_diffs = all_diffs_result.scalars().all()

    all_resolved = all(d.resolved == 1 for d in all_diffs)
    if all_resolved:
        invoice.status = InvoiceStatus.CONFIRMED

    # Audit log for diff resolution
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="parsing_diff",
        entity_id=diff_id,
        action="resolve",
        old_value=old_diff_value,
        new_value={
            "field_name": field_name,
            "final_value": final_value,
            "source": resolve_request.source,
            "resolved": 1,
            "invoice_id": invoice_id,
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()

    return {
        "message": "差异已解决",
        "field_name": field_name,
        "final_value": final_value,
        "all_resolved": all_resolved
    }


@router.post("/{invoice_id}/confirm")
async def confirm_invoice(
    invoice_id: int,
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """确认发票，标记所有差异为已解决。"""
    # Get invoice
    invoice_query = select(Invoice).where(Invoice.id == invoice_id)
    invoice_result = await db.execute(invoice_query)
    invoice = invoice_result.scalar_one_or_none()

    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")

    critical_fields = [
        "invoice_number",
        "issue_date",
        "total_with_tax",
        "buyer_name",
        "buyer_tax_id",
        "seller_name",
        "seller_tax_id",
        "item_name",
    ]
    missing_fields = [field for field in critical_fields if not getattr(invoice, field)]
    if missing_fields:
        raise HTTPException(
            status_code=400,
            detail="无法确认：必填字段缺失，请补全后再确认。"
        )

    # Check if LLM result exists
    llm_query = select(LlmResult).where(LlmResult.invoice_id == invoice_id)
    llm_result = await db.execute(llm_query)
    llm = llm_result.scalar_one_or_none()
    has_llm = llm is not None

    # Get parsing diffs to verify comparison exists
    diffs_query = select(ParsingDiff).where(ParsingDiff.invoice_id == invoice_id)
    diffs_result = await db.execute(diffs_query)
    diffs = diffs_result.scalars().all()

    if has_llm and not diffs:
        raise HTTPException(
            status_code=400,
            detail="无法确认：缺少OCR和LLM的比对结果。请重新解析发票。"
        )

    # Mark all diffs as resolved
    for diff in diffs:
        if diff.resolved == 0:
            # Use existing final_value or ocr_value as default
            if not diff.final_value:
                diff.final_value = diff.ocr_value or diff.llm_value
            diff.resolved = 1

    # Update invoice status
    old_status = invoice.status.value if invoice.status else None
    invoice.status = InvoiceStatus.CONFIRMED

    # Audit log for confirmation
    client_info = get_client_info(request)
    await log_audit_no_commit(
        db=db,
        entity_type="invoice",
        entity_id=invoice_id,
        action="confirm",
        old_value={"status": old_status},
        new_value={
            "status": InvoiceStatus.CONFIRMED.value,
            "resolved_diffs": len(diffs),
            "source": "llm" if has_llm else "ocr",
        },
        ip_address=client_info.get("ip_address"),
        user_agent=client_info.get("user_agent"),
    )

    await db.commit()

    return {"message": "发票已确认", "resolved_count": len(diffs)}


@router.get("/export/csv")
@limiter.limit("10/minute")
async def export_invoices_csv(
    request: Request,
    invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
    status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
    owner: Optional[str] = Query(None, description="归属人筛选"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db)
):
    """导出发票为CSV格式"""
    import csv
    from urllib.parse import quote

    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    start_date_obj = _parse_date_param(start_date, "start_date")
    if start_date_obj:
        query = query.where(Invoice.issue_date >= start_date_obj)
    end_date_obj = _parse_date_param(end_date, "end_date")
    if end_date_obj:
        query = query.where(Invoice.issue_date <= end_date_obj)

    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()

    # Create CSV content
    output = BytesIO()
    # Write BOM for Excel compatibility with Chinese characters
    output.write(b'\xef\xbb\xbf')

    import codecs
    writer = csv.writer(codecs.getwriter('utf-8')(output))

    # Header row
    writer.writerow([
        '发票号码', '开票日期', '购买方名称', '购买方纳税人识别号',
        '销售方名称', '销售方纳税人识别号', '项目名称',
        '金额', '税额', '价税合计', '税率',
        '状态', '归属人', '文件名', '创建时间'
    ])

    # Data rows
    for inv in invoices:
        writer.writerow([
            inv.invoice_number or '',
            str(inv.issue_date) if inv.issue_date else '',
            inv.buyer_name or '',
            inv.buyer_tax_id or '',
            inv.seller_name or '',
            inv.seller_tax_id or '',
            inv.item_name or '',
            str(inv.amount) if inv.amount else '',
            str(inv.tax_amount) if inv.tax_amount else '',
            str(inv.total_with_tax) if inv.total_with_tax else '',
            inv.tax_rate or '',
            inv.status.value if inv.status else '',
            inv.owner or '',
            inv.file_name or '',
            inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else ''
        ])

    output.seek(0)

    filename = quote('发票导出.csv')
    return StreamingResponse(
        output,
        media_type='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )


@router.get("/export/excel")
@limiter.limit("10/minute")
async def export_invoices_excel(
    request: Request,
    invoice_ids: Optional[str] = Query(None, description="发票ID列表，逗号分隔"),
    status: Optional[InvoiceStatus] = Query(None, description="状态筛选"),
    owner: Optional[str] = Query(None, description="归属人筛选"),
    start_date: Optional[str] = Query(None, description="开始日期"),
    end_date: Optional[str] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db)
):
    """导出发票为Excel格式"""
    from urllib.parse import quote

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel导出需要安装openpyxl库")

    query = select(Invoice)

    ids = _parse_invoice_ids(invoice_ids)
    if ids:
        query = query.where(Invoice.id.in_(ids))
    if status:
        query = query.where(Invoice.status == status)
    if owner:
        query = query.where(Invoice.owner == owner)
    start_date_obj = _parse_date_param(start_date, "start_date")
    if start_date_obj:
        query = query.where(Invoice.issue_date >= start_date_obj)
    end_date_obj = _parse_date_param(end_date, "end_date")
    if end_date_obj:
        query = query.where(Invoice.issue_date <= end_date_obj)

    query = query.order_by(Invoice.created_at.desc())
    result = await db.execute(query)
    invoices = result.scalars().all()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票列表"

    # Header row
    headers = [
        '发票号码', '开票日期', '购买方名称', '购买方纳税人识别号',
        '销售方名称', '销售方纳税人识别号', '项目名称',
        '金额', '税额', '价税合计', '税率',
        '状态', '归属人', '文件名', '创建时间'
    ]

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row, inv in enumerate(invoices, 2):
        data = [
            inv.invoice_number or '',
            str(inv.issue_date) if inv.issue_date else '',
            inv.buyer_name or '',
            inv.buyer_tax_id or '',
            inv.seller_name or '',
            inv.seller_tax_id or '',
            inv.item_name or '',
            float(inv.amount) if inv.amount else '',
            float(inv.tax_amount) if inv.tax_amount else '',
            float(inv.total_with_tax) if inv.total_with_tax else '',
            inv.tax_rate or '',
            inv.status.value if inv.status else '',
            inv.owner or '',
            inv.file_name or '',
            inv.created_at.strftime('%Y-%m-%d %H:%M:%S') if inv.created_at else ''
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = quote('发票导出.xlsx')
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{filename}"
        }
    )
